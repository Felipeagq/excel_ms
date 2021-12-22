# Importamos librerias
import os 
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import json


# Cargamos el json
with open('excel.json') as file:
    creation = json.load(file)


# Creamos el objeto de Excel
wb = Workbook()

# verificamos que exista el documento de Excel
wb = Workbook()
wb.create_sheet(creation["name_sheet"],creation["position_sheet"])
ws = wb[creation["name_sheet"]]
row = 1
column = 1
if creation["orientation"] == "y":
    print("llenado de forma vertical")
for data in  creation["data"]:
    for values in data.values():
        ws.cell(row=row,column=column,value=list(data.keys())[0])
        if creation["orientation"] == "y":
            row = row +1
        else:
            column = column + 1
        for value in values:
            print(value)
            ws.cell(row=row,column=column,value=value)
            if creation["orientation"] == "y":
                row = row + 1
            else:
                column = column + 1
        if creation["orientation"] == "y":
            row = 1
            column = column + 1
        else:
            column = 1
            row = row + 1
            
ws.cell
wb.save(creation["name_document"])






